"""
RAG infrastructure: document loading, semantic chunking, FAISS index management,
live web fetching (via MCP), and Tavily search.

All index operations are incremental — a rebuild is triggered only when source
files change (mtime/size fingerprint) or RagConfig parameters differ.
"""

from __future__ import annotations

import asyncio
import json
import os
import time
from dataclasses import dataclass
from typing import Any, Dict, List, Optional
from urllib.parse import urlparse

from bs4 import BeautifulSoup
from langchain_core.documents import Document
from langchain_experimental.text_splitter import SemanticChunker
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_openai import OpenAIEmbeddings
import openpyxl
from langchain_community.document_loaders import PyPDFLoader
from langchain_community.vectorstores import FAISS


# -------------------------
# Config
# -------------------------

@dataclass
class RagConfig:
    """Chunking and embedding settings; also used as the index cache key."""

    chunk_size: int = 1500
    chunk_overlap: int = 200
    embeddings_model: str = "text-embedding-3-small"
    semantic_chunking: bool = True
    semantic_breakpoint_threshold_type: str = "percentile"  # "percentile" | "standard_deviation" | "interquartile"
    semantic_breakpoint_threshold_amount: float = 95.0


# -------------------------
# Basic helpers
# -------------------------

def ensure_dir(path: str) -> None:
    """Create directory (and parents) if it does not already exist."""
    os.makedirs(path, exist_ok=True)


def list_files(folder: str, extensions: tuple[str, ...]) -> List[str]:
    """Return sorted absolute paths of all files with matching extensions, recursively."""
    files: List[str] = []
    for root, _, names in os.walk(folder):
        for name in names:
            if name.lower().endswith(tuple(ext.lower() for ext in extensions)):
                files.append(os.path.join(root, name))
    return sorted(files)


def read_json(path: str) -> Dict[str, Any]:
    """Load and return a JSON file as a dict."""
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def write_json(path: str, data: Dict[str, Any]) -> None:
    """Serialise data to JSON, creating parent directories if needed."""
    ensure_dir(os.path.dirname(path) or ".")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def append_jsonl(path: str, obj: dict) -> None:
    """Append one JSON object as a newline-delimited record; creates file and dirs if missing."""
    folder = os.path.dirname(path) or "."
    ensure_dir(folder)
    with open(path, "a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")


def collect_fingerprints(folder: str, extensions: tuple[str, ...]) -> List[Dict[str, Any]]:
    """Return mtime + size fingerprints for matching files; used as the index cache-invalidation key.

    Paths are stored relative to folder so fingerprints stay valid across OS/Docker boundaries.
    """
    fingerprints = []
    abs_folder = os.path.abspath(folder)
    for path in list_files(folder, extensions):
        st = os.stat(path)
        rel = os.path.relpath(os.path.abspath(path), abs_folder)
        fingerprints.append({
            "path": rel,
            "mtime": int(st.st_mtime),
            "size": int(st.st_size),
        })
    return fingerprints


# -------------------------
# HTML parsing
# -------------------------

def html_string_to_text(html: str) -> str:
    """Extract readable text from HTML, preferring main content containers and deduplicating lines."""
    soup = BeautifulSoup(html, "lxml")

    for tag in soup(["script", "style", "noscript", "svg", "form", "button", "header", "footer", "nav", "aside"]):
        tag.decompose()

    selectors = [
        "main article",
        "main .editor-content",
        "main .article-area__article",
        "main .journal-content-article",
        "main .portlet-content",
        "main",
        "article",
        '[role="main"]',
    ]

    container = None
    for sel in selectors:
        node = soup.select_one(sel)
        if node is not None:
            container = node
            break

    if container is None:
        container = soup.body if soup.body is not None else soup

    text = container.get_text(separator="\n")

    lines = [ln.strip() for ln in text.splitlines()]
    lines = [
        ln for ln in lines
        if ln
        and ln.lower() not in {
            "skip to content",
            "go to the footer section",
            "menu",
            "search",
            "home",
            "kontakt",
            "zaloguj do pue/ezus",
            "zarejestruj w pue/ezus",
        }
        and len(ln) > 1
    ]

    cleaned: List[str] = []
    seen = set()
    for ln in lines:
        key = ln.lower()
        if key not in seen:
            cleaned.append(ln)
            seen.add(key)

    return "\n".join(cleaned)


def html_to_text(html_path: str) -> str:
    """Read an HTML file from disk and return its cleaned text content."""
    with open(html_path, "r", encoding="utf-8", errors="ignore") as f:
        return html_string_to_text(f.read())


# -------------------------
# Document loading
# -------------------------

def load_pdf_documents(pdf_paths: List[str]) -> List[Any]:
    """Load all PDFs via PyPDFLoader and return a flat list of LangChain Documents."""
    docs: List[Any] = []
    for path in pdf_paths:
        docs.extend(PyPDFLoader(path).load())
    return docs


def load_xlsx_documents(xlsx_paths: List[str]) -> List[Document]:
    """Load each worksheet as a tab-delimited Document; preserves the header row as the first line."""
    docs: List[Document] = []
    for path in xlsx_paths:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header = [str(c) if c is not None else "" for c in rows[0]]
            lines = ["\t".join(header)]
            for row in rows[1:]:
                lines.append("\t".join("" if c is None else str(c) for c in row))
            content = "\n".join(lines)
            docs.append(
                Document(
                    page_content=content,
                    metadata={
                        "source": os.path.abspath(path),
                        "page": sheet_name,
                        "file_type": "xlsx",
                    },
                )
            )
        wb.close()
    return docs


def load_html_documents(html_paths: List[str]) -> List[Document]:
    """Load HTML files and return one Document per file with cleaned text content."""
    docs: List[Document] = []
    for path in html_paths:
        docs.append(
            Document(
                page_content=html_to_text(path),
                metadata={
                    "source": os.path.abspath(path),
                    "page": None,
                    "file_type": "html",
                },
            )
        )
    return docs


# -------------------------
# Chunking
# -------------------------

def split_into_chunks(docs: List[Any], cfg: RagConfig) -> List[Any]:
    """Semantic split followed by a size guard to prevent oversized chunks from reaching the embedder."""
    if cfg.semantic_chunking:
        embeddings = OpenAIEmbeddings(model=cfg.embeddings_model, chunk_size=64)
        semantic_splitter = SemanticChunker(
            embeddings,
            breakpoint_threshold_type=cfg.semantic_breakpoint_threshold_type,
            breakpoint_threshold_amount=cfg.semantic_breakpoint_threshold_amount,
        )
        # Guard: split any oversized chunks with the character splitter
        size_splitter = RecursiveCharacterTextSplitter(
            chunk_size=cfg.chunk_size,
            chunk_overlap=cfg.chunk_overlap,
            separators=["\n\n", "\n", " ", ""],
        )
        semantic_chunks = semantic_splitter.split_documents(docs)
        chunks = size_splitter.split_documents(semantic_chunks)
    else:
        splitter = RecursiveCharacterTextSplitter(
            chunk_size=cfg.chunk_size,
            chunk_overlap=cfg.chunk_overlap,
            separators=["\n\n", "\n", " ", ""],
        )
        chunks = splitter.split_documents(docs)

    for i, chunk in enumerate(chunks):
        chunk.metadata = dict(chunk.metadata or {})
        chunk.metadata["chunk_id"] = i

    return chunks


# -------------------------
# Indexing
# -------------------------

def build_index_from_folder(
    books_folder: str,
    index_folder: str,
    cfg: RagConfig,
) -> Dict[str, Any]:
    """Build a FAISS index from all PDF/HTML/XLSX in the folder and write a manifest for cache validation."""
    t0 = time.time()

    pdfs = list_files(books_folder, (".pdf",))
    htmls = list_files(books_folder, (".html",))
    xlsxs = list_files(books_folder, (".xlsx",))

    if not pdfs and not htmls and not xlsxs:
        raise RuntimeError(f"No PDF/HTML/XLSX files found in: {books_folder}")

    docs: List[Any] = []
    if pdfs:
        docs.extend(load_pdf_documents(pdfs))
    if htmls:
        docs.extend(load_html_documents(htmls))
    if xlsxs:
        docs.extend(load_xlsx_documents(xlsxs))

    chunks = split_into_chunks(docs, cfg)

    embeddings = OpenAIEmbeddings(model=cfg.embeddings_model, chunk_size=64)
    vs = FAISS.from_documents(chunks, embeddings)

    ensure_dir(index_folder)
    vs.save_local(index_folder)

    manifest = {
        "books_folder": os.path.abspath(books_folder),
        "index_folder": os.path.abspath(index_folder),
        "built_at_unix": int(time.time()),
        "build_seconds": round(time.time() - t0, 3),
        "cfg": cfg.__dict__,
        "inputs": {
            "pdf": collect_fingerprints(books_folder, (".pdf",)),
            "html": collect_fingerprints(books_folder, (".html",)),
            "xlsx": collect_fingerprints(books_folder, (".xlsx",)),
        },
        "n_docs": len(docs),
        "n_chunks": len(chunks),
    }

    write_json(os.path.join(index_folder, "manifest.json"), manifest)
    return manifest


def is_index_up_to_date(books_folder: str, index_folder: str, cfg: Optional[RagConfig] = None) -> bool:
    """Return True if the saved index matches current source files and config; False triggers a rebuild."""
    manifest_path = os.path.join(index_folder, "manifest.json")
    if not os.path.exists(manifest_path):
        return False

    old = read_json(manifest_path)

    if cfg is not None and old.get("cfg") != cfg.__dict__:
        return False

    old_inputs = old.get("inputs", {})
    current_pdf = collect_fingerprints(books_folder, (".pdf",))
    current_html = collect_fingerprints(books_folder, (".html",))
    current_xlsx = collect_fingerprints(books_folder, (".xlsx",))

    return (
        old_inputs.get("pdf", []) == current_pdf
        and old_inputs.get("html", []) == current_html
        and old_inputs.get("xlsx", []) == current_xlsx
    )


def load_index(index_folder: str, cfg: RagConfig) -> FAISS:
    """Load a previously saved FAISS index from disk."""
    embeddings = OpenAIEmbeddings(model=cfg.embeddings_model, chunk_size=64)
    return FAISS.load_local(
        index_folder,
        embeddings,
        allow_dangerous_deserialization=True,
    )


def build_or_load_index(books_folder: str, index_folder: str, cfg: RagConfig) -> FAISS:
    """Return a ready-to-query FAISS index, rebuilding only when sources or config changed."""
    if not is_index_up_to_date(books_folder, index_folder, cfg):
        build_index_from_folder(books_folder, index_folder, cfg)
    return load_index(index_folder, cfg)


def retrieve_chunks(vs: FAISS, query: str, k: int = 4) -> List[Any]:
    """Return the top-k most similar document chunks from the vectorstore."""
    return vs.similarity_search(query, k=k)


# -------------------------
# Tavily search
# -------------------------

def tavily_search(
    query: str,
    api_key: str,
    max_results: int = 4,
    allowed_domains: Optional[List[str]] = None,
) -> List[Dict[str, Any]]:
    """Search via Tavily; optionally restrict results to allowed_domains."""
    from tavily import TavilyClient
    client = TavilyClient(api_key=api_key)
    kwargs: Dict[str, Any] = {"max_results": max_results}
    if allowed_domains:
        kwargs["include_domains"] = allowed_domains
    response = client.search(query, **kwargs)
    results = response.get("results", [])
    if allowed_domains:
        results = [r for r in results if is_allowed_domain(str(r.get("url", "")), allowed_domains)]
    return results


# -------------------------
# Browser / web fetch
# -------------------------

def _domain(url: str) -> str:
    """Extract the lowercase hostname from a URL."""
    return (urlparse(url).hostname or "").lower()


def is_allowed_domain(url: str, allowed_domains: List[str]) -> bool:
    """Return True if the URL's hostname matches or is a subdomain of any entry in allowed_domains."""
    host = _domain(url)
    return any(host == d or host.endswith("." + d) for d in allowed_domains)


async def _mcp_fetch_async(url: str, timeout_s: float = 20.0) -> str:
    """Spin up a local MCP stdio server and call its fetch tool for the given URL."""
    from mcp import ClientSession, StdioServerParameters
    from mcp.client.stdio import stdio_client

    server_params = StdioServerParameters(
        command="python",
        args=["-m", "mcp_server_fetch"],
    )
    async with stdio_client(server_params) as (read, write):
        async with ClientSession(read, write) as session:
            await asyncio.wait_for(session.initialize(), timeout=timeout_s)
            result = await asyncio.wait_for(
                session.call_tool("fetch", {"url": url, "max_length": 50000}),
                timeout=timeout_s,
            )
            if result.content:
                return result.content[0].text or ""
            return ""


def mcp_fetch_text(
    url: str,
    allowed_domains: Optional[List[str]] = None,
    timeout_ms: int = 20000,
) -> str:
    """Synchronous wrapper around _mcp_fetch_async; raises ValueError if domain is not allowed."""
    if allowed_domains is not None and not is_allowed_domain(url, allowed_domains):
        raise ValueError(f"Domain not allowed: {url}")
    try:
        return asyncio.run(_mcp_fetch_async(url, timeout_s=timeout_ms / 1000))
    except (asyncio.TimeoutError, asyncio.CancelledError, KeyboardInterrupt, Exception) as e:
        print(f"[MCP fetch] timeout/error for {url}: {repr(e)}")
        return ""


def mcp_fetch_to_file(
    url: str,
    out_dir: str,
    allowed_domains: Optional[List[str]] = None,
    timeout_ms: int = 20000,
) -> Dict[str, str]:
    """Fetch URL via MCP, write raw text to out_dir, return {'txt_path', 'text'}."""
    ensure_dir(out_dir)

    safe_name = url.replace("https://", "").replace("http://", "")
    safe_name = safe_name.replace("/", "_").replace("?", "_").replace("&", "_")[:180]

    text = mcp_fetch_text(url, allowed_domains=allowed_domains, timeout_ms=timeout_ms)

    txt_path = os.path.join(out_dir, f"{safe_name}.txt")
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write(text)

    return {"txt_path": os.path.abspath(txt_path), "text": text}
